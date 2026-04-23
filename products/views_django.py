from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from decimal import Decimal
import csv
import io
from products.models import Product, Category, Supplier
from products.forms import ProductForm


def is_admin(user):
    return user.is_superuser or user.role == 'admin'


def is_staff_user(user):
    return user.is_superuser or user.role == 'admin' or user.is_staff


@login_required
def product_list(request):
    products = Product.objects.select_related('category', 'supplier').order_by('-created_at')
    
    search = request.GET.get('search')
    if search:
        products = products.filter(name__icontains=search) | products.filter(sku__icontains=search)
    
    stock_filter = request.GET.get('stock')
    if stock_filter == 'low':
        products = products.filter(stock_quantity__lte=10, is_active=True)
    elif stock_filter == 'out':
        products = products.filter(stock_quantity=0, is_active=True)
    elif stock_filter == 'available':
        products = products.filter(stock_quantity__gt=0, is_active=True)
    
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    
    categories = Category.objects.all()
    
    paginator = Paginator(products, 20)
    page = request.GET.get('page')
    products = paginator.get_page(page)
    return render(request, 'products/product_list.html', {
        'products': products,
        'categories': categories,
        'filters': {
            'search': search,
            'stock': stock_filter,
            'category': category_id
        }
    })


@login_required
def product_create(request):
    if not is_staff_user(request.user):
        messages.warning(request, 'You do not have permission to add products.')
        return redirect('products:product_list')
    
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product created successfully')
            return redirect('products:product_list')
    else:
        form = ProductForm()
    return render(request, 'products/product_form.html', {'form': form})


@login_required
def product_edit(request, pk):
    if not is_staff_user(request.user):
        messages.warning(request, 'You do not have permission to edit products.')
        return redirect('products:product_list')
    
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully')
            return redirect('products:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'products/product_form.html', {'form': form, 'product': product})


@login_required
def product_delete(request, pk):
    if not is_staff_user(request.user):
        messages.warning(request, 'You do not have permission to delete products.')
        return redirect('products:product_list')
    
    product = get_object_or_404(Product, pk=pk)
    product.delete()
    messages.success(request, 'Product deleted successfully')
    return redirect('products:product_list')


@login_required
def product_bulk_upload(request):
    if not is_staff_user(request.user):
        messages.warning(request, 'You do not have permission to upload products.')
        return redirect('products:product_list')

    template_headers = [
        'name', 'sku', 'barcode', 'category', 'supplier', 'description',
        'cost_price', 'selling_price', 'stock_quantity', 'low_stock_threshold', 'is_active'
    ]
    upload_errors = []
    created_count = 0
    updated_count = 0

    if request.method == 'POST':
        uploaded_file = request.FILES.get('product_file')
        if not uploaded_file:
            messages.error(request, 'Please select a CSV or XLSX file to upload.')
        else:
            # Validate file type
            if uploaded_file.content_type not in ['text/csv', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
                messages.error(request, 'Invalid file type. Only CSV and XLSX files are allowed.')
                return redirect('products:product_bulk_upload')
            
            # Validate file size (5MB limit)
            if uploaded_file.size > 5 * 1024 * 1024:
                messages.error(request, 'File size exceeds 5MB limit.')
                return redirect('products:product_bulk_upload')
            
            file_name = uploaded_file.name.lower()
            try:
                if file_name.endswith('.csv'):
                    decoded = uploaded_file.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(decoded))
                    rows = list(reader)
                elif file_name.endswith('.xlsx'):
                    from openpyxl import load_workbook
                    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
                    sheet = workbook.active
                    rows = []
                    header_row = None
                    for row in sheet.iter_rows(values_only=True):
                        if header_row is None:
                            header_row = [str(cell).strip() if cell is not None else '' for cell in row]
                            continue
                        if not any(cell not in (None, '') for cell in row):
                            continue
                        rows.append({header_row[i]: row[i] for i in range(min(len(header_row), len(row)))})
                else:
                    raise ValueError('Unsupported file format. Please upload a CSV or XLSX file.')

                if not rows:
                    raise ValueError('The uploaded file has no data rows.')

                for index, row in enumerate(rows, start=2):
                    sku = str(row.get('sku') or '').strip()
                    name = str(row.get('name') or '').strip()
                    if not sku or not name:
                        upload_errors.append(f'Row {index}: name and sku are required.')
                        continue

                    category_name = str(row.get('category') or '').strip()
                    supplier_name = str(row.get('supplier') or '').strip()
                    category = None
                    supplier = None
                    if category_name:
                        category, _ = Category.objects.get_or_create(name=category_name)
                    if supplier_name:
                        supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

                    def parse_decimal(value, default='0'):
                        if value in (None, ''):
                            return Decimal(default)
                        try:
                            return Decimal(str(value).strip())
                        except Exception:
                            return Decimal(default)

                    def parse_int(value, default=0):
                        if value in (None, ''):
                            return default
                        try:
                            return int(float(value))
                        except Exception:
                            return default

                    def parse_bool(value):
                        if value in (None, ''):
                            return True
                        return str(value).strip().lower() in ('1', 'true', 'yes', 'y', 'active')

                    defaults = {
                        'name': name,
                        'barcode': str(row.get('barcode') or '').strip(),
                        'category': category,
                        'supplier': supplier,
                        'description': str(row.get('description') or '').strip(),
                        'cost_price': parse_decimal(row.get('cost_price')),
                        'selling_price': parse_decimal(row.get('selling_price')),
                        'stock_quantity': parse_int(row.get('stock_quantity')),
                        'low_stock_threshold': parse_int(row.get('low_stock_threshold'), 10),
                        'is_active': parse_bool(row.get('is_active')),
                    }

                    _, created = Product.objects.update_or_create(sku=sku, defaults=defaults)
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                if upload_errors:
                    messages.warning(request, f'Upload completed with {len(upload_errors)} warnings. Created {created_count}, updated {updated_count}.')
                else:
                    messages.success(request, f'Upload completed successfully. Created {created_count}, updated {updated_count}.')
            except Exception as exc:
                messages.error(request, str(exc))

    return render(request, 'products/product_bulk_upload.html', {
        'template_headers': template_headers,
        'upload_errors': upload_errors,
        'created_count': created_count,
        'updated_count': updated_count,
    })


@login_required
def download_product_upload_template(request):
    if not is_staff_user(request.user):
        messages.warning(request, 'You do not have permission to download templates.')
        return redirect('products:product_list')

    headers = [
        'name', 'sku', 'barcode', 'category', 'supplier', 'description',
        'cost_price', 'selling_price', 'stock_quantity', 'low_stock_threshold', 'is_active'
    ]
    sample_row = [
        'Example Product', 'EXT-001', '1234567890123', 'General', 'Default Supplier',
        'Sample product description', '100.00', '150.00', '50', '10', 'True'
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_row)

    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product-upload-template.csv"'
    return response


@login_required
def stock_history(request):
    from products.models import StockHistory
    from django.db.models import Q
    
    history = StockHistory.objects.select_related('product', 'created_by').order_by('-created_at')
    
    product_id = request.GET.get('product')
    action = request.GET.get('action')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if product_id:
        history = history.filter(product_id=product_id)
    if action:
        history = history.filter(action=action)
    if date_from:
        history = history.filter(created_at__date__gte=date_from)
    if date_to:
        history = history.filter(created_at__date__lte=date_to)
    
    products = Product.objects.filter(is_active=True)
    
    paginator = Paginator(history, 50)
    page = request.GET.get('page')
    history = paginator.get_page(page)
    
    return render(request, 'products/stock_history.html', {
        'history': history,
        'products': products,
        'filters': {
            'product': product_id,
            'action': action,
            'date_from': date_from,
            'date_to': date_to
        }
    })


@login_required
def category_list(request):
    from products.models import Category
    categories = Category.objects.all().order_by('name')
    paginator = Paginator(categories, 20)
    page = request.GET.get('page')
    categories = paginator.get_page(page)
    return render(request, 'products/category_list.html', {'categories': categories})


@login_required
def category_create(request):
    from products.models import Category
    if not is_staff_user(request.user):
        messages.warning(request, 'Permission denied.')
        return redirect('products:category_list')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        Category.objects.create(name=name, description=description)
        messages.success(request, 'Category created.')
        return redirect('products:category_list')
    return render(request, 'products/category_form.html')


@login_required
def category_delete(request, pk):
    from products.models import Category
    if not is_staff_user(request.user):
        messages.warning(request, 'Permission denied.')
        return redirect('products:category_list')
    
    category = get_object_or_404(Category, pk=pk)
    category.delete()
    messages.success(request, 'Category deleted.')
    return redirect('products:category_list')


@login_required
def supplier_list(request):
    from products.models import Supplier
    suppliers = Supplier.objects.all().order_by('name')
    paginator = Paginator(suppliers, 20)
    page = request.GET.get('page')
    suppliers = paginator.get_page(page)
    return render(request, 'products/supplier_list.html', {'suppliers': suppliers})


@login_required
def supplier_create(request):
    from products.models import Supplier
    if not is_staff_user(request.user):
        messages.warning(request, 'Permission denied.')
        return redirect('products:supplier_list')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        contact_person = request.POST.get('contact_person')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        Supplier.objects.create(name=name, contact_person=contact_person, email=email, phone=phone, address=address)
        messages.success(request, 'Supplier created.')
        return redirect('products:supplier_list')
    return render(request, 'products/supplier_form.html')


@login_required
def supplier_delete(request, pk):
    from products.models import Supplier
    if not is_staff_user(request.user):
        messages.warning(request, 'Permission denied.')
        return redirect('products:supplier_list')
    
    supplier = get_object_or_404(Supplier, pk=pk)
    supplier.delete()
    messages.success(request, 'Supplier deleted.')
    return redirect('products:supplier_list')